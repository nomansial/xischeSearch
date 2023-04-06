import openpyxl
from openpyxl import Workbook, load_workbook


class tools:

    def getData(self):

        workbook = openpyxl.load_workbook("C:\Users\user\PycharmProjects\xischeSearch\keywords.xlsx")
        sheet = workbook["Tab"]
        totalrows = sheet.max_row
        totalcolumns = sheet.max_column

        print(totalcolumns)

        mainlist = []

        for i in range(1, totalrows + 1):
            datalist = []
            for j in range(1, totalcolumns + 1):
                data = sheet.cell(row=i, column=j).value
                datalist.insert(j, data)

            mainlist.insert(i, datalist)
            print(mainlist)
        return mainlist
